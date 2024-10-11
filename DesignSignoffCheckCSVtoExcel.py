#!/bin/env python3.7

import re
import sys
import csv
import openpyxl

def ChangeColumnWidth( sheet, column, width ):

  columnLetter = openpyxl.utils.get_column_letter( column )
  sheet.column_dimensions[ columnLetter ].width = width

def ChangeColumnWidthWithMaxStringLength( sheet, start_row, end_row, start_column, end_column ):

  for column in range( start_column, end_column + 1 ):
    max_width = 0

    for row in range( start_row, end_row + 1 ):
      width = 0
      value = sheet.cell( row, column ).value

      if not value is None:
        width = len( value )
      else:
        width = 0

      if width > max_width:
        max_width = width

    ChangeColumnWidth( sheet, column, max_width )

def MergeCells( sheet, startRow, endRow, startColumn, endColumn ):

  sheet.merge_cells( start_row = startRow, end_row = endRow, start_column = startColumn, end_column = endColumn )

def CreateSheet( workBook, csvFile ):

  sheet = workBook.create_sheet( re.sub( r'\.csv$', "", csvFile ) )

  read = open( csvFile )

  csvRead = csv.reader( read )
  
  row = 1
  column = 1

  headRow = list()
  headRow.append( 1 )

  for line in csvRead:
    if len( line ) > 0:
      for token in line:
        sheet.cell( row, column ).font = gFontCalibri
        sheet.cell( row, column ).value = token
        sheet.cell( row, column ).border = gThinBorder
        column += 1

      sheet.cell( row, 1 ).alignment = gAlignCenter
      sheet.cell( row, 2 ).alignment = gAlignCenter

    else:
      headRow.append( row + 1 )

    row+= 1
    column = 1

  read.close()

  for row in headRow:
    for column in range( 1, 7 ):
      sheet.cell( row, column ).font = gFontCalibriAndBold
      sheet.cell( row, column ).alignment = gAlignCenter
      sheet.cell( row, column ).fill = gBackgroundColorYellow

  ChangeColumnWidth( sheet, 1, 15 )
  ChangeColumnWidth( sheet, 2, 12 )
  ChangeColumnWidth( sheet, 3, 35 )
  ChangeColumnWidth( sheet, 4, 5 )
  ChangeColumnWidth( sheet, 5, 15 )
  ChangeColumnWidth( sheet, 6, 60 )

def SummarySheet( workBook, csvFile ):

  sheet = workBook.active
  sheet.title = re.sub( r'\.csv$', "", csvFile )

  read = open( csvFile, "r" )

  csvRead = csv.reader( read )

  row = 1
  column = 1

  for line in csvRead:
    for tokens in line:
      sheet.cell( row, column ).value = tokens
      sheet.cell( row, column ).font = gFontCalibriAndBold
      column += 1

    row += 1
    column = 1

  read.close()

  ChangeColumnWidth( sheet, 1, 12 )
  ChangeColumnWidth( sheet, 2, 40 )
  ChangeColumnWidth( sheet, 3, 13 )
  ChangeColumnWidth( sheet, 4, 13 )
  ChangeColumnWidth( sheet, 5, 15 )
  ChangeColumnWidth( sheet, 6, 9 )
  ChangeColumnWidth( sheet, 7, 100 )

  MergeCells( sheet, 1, 2, 1, 1 )
  MergeCells( sheet, 1, 2, 2, 2 )
  MergeCells( sheet, 1, 2, 6, 7 )

  for row in range( 1, int( sheet.max_row / 2 ) ):
    row = row * 2 + 1
    for column in range( 1, 5 ):
      MergeCells( sheet, row, row + 1, column, column )
      sheet.cell( row, column ).alignment = gAlignCenter

  for column in range( 1, int( sheet.max_column ) ):
    sheet.cell( 1, column ).alignment = gAlignCenter
    sheet.cell( 2, column ).alignment = gAlignCenter

  ChangeColumnWidthWithMaxStringLength( sheet, 1, sheet.max_row, 1, sheet.max_column - 2 )
  ChangeColumnWidthWithMaxStringLength( sheet, 3, sheet.max_row, sheet.max_column - 1, sheet.max_column )

  for column in range( 3, 5 ):
    for row in range( 1, int( sheet.max_row / 2 ) ):

      row = row * 2 + 1

      value = sheet.cell( row, column ).value

      if value == "PASSED":
        sheet.cell( row, column ).font = gFontGreenCalibriAndBold
        sheet.cell( row, column + 1 ).font = gFontGreenCalibriAndBold
        sheet.cell( row + 1, column + 1 ).font = gFontGreenCalibriAndBold

      elif value == "FAILED":
        sheet.cell( row, column ).font = gFontRedCalibriAndBold
        sheet.cell( row, column + 1 ).font = gFontRedCalibriAndBold
        sheet.cell( row + 1, column + 1 ).font = gFontRedCalibriAndBold

      elif value == "NONE":
        sheet.cell( row, column ).font = gFontBlackCalibriAndBold
        sheet.cell( row, column + 1 ).font = gFontBlackCalibriAndBold
        sheet.cell( row + 1, column + 1 ).font = gFontBlackCalibriAndBold

  sheet.cell( 1, 1 ).border = openpyxl.styles.borders.Border( left = gThick, right = gThin, top = gThick, bottom = gThin )
  sheet.cell( 2, 1 ).border = openpyxl.styles.borders.Border( left = gThick, right = gThin, top = gThick, bottom = gThin )

  for row in range( 3, sheet.max_row ):
    sheet.cell( row, 1 ).border = openpyxl.styles.borders.Border( left = gThick, right = gThin )

  sheet.cell( sheet.max_row, 1 ).border = openpyxl.styles.borders.Border( left = gThick, right = gThin, bottom = gThick )

  for column in range( 2, sheet.max_column ):
    sheet.cell( 1, column ).border = openpyxl.styles.borders.Border( top = gThick )
    sheet.cell( 2, column ).border = openpyxl.styles.borders.Border( bottom = gThin )
    sheet.cell( sheet.max_row, column ).border = openpyxl.styles.borders.Border( bottom = gThick )

  for row in range( 3, sheet.max_row ):
    sheet.cell( row, sheet.max_column ).border = openpyxl.styles.borders.Border( right = gThick )

  sheet.cell( 1, sheet.max_column ).border = openpyxl.styles.borders.Border( right = gThick, top = gThick )
  sheet.cell( 2, sheet.max_column ).border = openpyxl.styles.borders.Border( right = gThick, bottom = gThin )
  sheet.cell( sheet.max_row, sheet.max_column ).border = openpyxl.styles.borders.Border( right = gThick, bottom = gThick )

gBackgroundColorYellow = openpyxl.styles.PatternFill( start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid" )

gFontCalibri = openpyxl.styles.Font( name = "Calibri" )
gFontCalibriAndBold = openpyxl.styles.Font( name = "Calibri", bold = True )
gFontRedCalibriAndBold = openpyxl.styles.Font( name = "Calibri", bold = True, color = "FF0000" )
gFontBlackCalibriAndBold = openpyxl.styles.Font( name = "Calibri", bold = True, color = "000000" )
gFontGreenCalibriAndBold = openpyxl.styles.Font( name = "Calibri", bold = True, color = "00A050" )

gThin = openpyxl.styles.borders.Side( border_style = openpyxl.styles.borders.BORDER_THIN )
gThick = openpyxl.styles.borders.Side( border_style = openpyxl.styles.borders.BORDER_THICK )
gThinBorder = openpyxl.styles.borders.Border( left = gThin, right = gThin, top = gThin, bottom = gThin )

gAlignRight = openpyxl.styles.Alignment( horizontal = "right", vertical = "center" )
gAlignCenter = openpyxl.styles.Alignment( horizontal = "center", vertical = "center" )

if __name__ == "__main__":

  workBook = openpyxl.Workbook()

  SummarySheet( workBook, sys.argv[ 2 ] )

  for csvFile in sys.argv[ 3: ]:
    CreateSheet( workBook, csvFile )

  workBook.save( sys.argv[ 1 ] )

  sys.exit( 0 )
